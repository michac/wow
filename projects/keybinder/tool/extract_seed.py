#!/usr/bin/env python3
"""ARCHIVAL importer — NOT part of the routine workflow.

The Bellular "Midnight Keybind Planner" workbook was the *original* seed source,
but it is now a **frozen archive**. The live source of truth is the hand-edited
`data/bellular-keybinds.seed.json` (it carries Tier-A/B corrections and the
modifier-grouped bar re-layout the workbook never had). The addon's `Data.lua`
is generated from that JSON by `tool/gen_data_lua.py` — that is the one step in
the live workflow.

This script exists only to *re-derive* the seed from the workbook for reference
/ diffing (e.g. if Bellular ships a new sheet worth comparing against). To avoid
clobbering the authoritative, corrected seed it writes to a **side file**
(`data/bellular-keybinds.archive-import.json`), never the canonical JSON, and it
does **not** write `Data.lua`. Diff the side file against the canonical seed by
hand and port anything worth keeping.

Usage:
    uv run --with openpyxl python tool/extract_seed.py [path-to.xlsx]
"""
import json
import sys
from pathlib import Path

import openpyxl

DEFAULT_XLSX = (
    "/mnt/c/Users/mchris/Downloads/"
    "Copy of Midnight Keybind Planner (Oldschool Edition).xlsx"
)
HERE = Path(__file__).resolve().parent
PROJECT = HERE.parent
# Side file only — the canonical seed (data/bellular-keybinds.seed.json) is
# hand-maintained and must not be overwritten by a workbook re-import.
JSON_OUT = PROJECT / "data" / "bellular-keybinds.archive-import.json"

# class tabs, in the workbook's tab order
CLASS_TABS = [
    "Death Knight", "Paladin", "Warrior", "Evoker", "Hunter", "Shaman",
    "Demon Hunter", "Druid", "Monk", "Rogue", "Mage", "Priest", "Warlock",
]


def cell(v):
    """Normalize a cell value to a trimmed string (ints stay int-like)."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def extract(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 1. bucket -> default keybind (+ the out-of-combat "bonus" binds)
    ws = wb["ENTER YOUR BINDS HERE"]
    buckets, bonus = [], []
    for r in range(3, ws.max_row + 1):
        cat, key = cell(ws.cell(r, 1).value), cell(ws.cell(r, 2).value)
        if cat:
            buckets.append({"category": cat, "keybind": key})
        bcat, bkey = cell(ws.cell(r, 3).value), cell(ws.cell(r, 4).value)
        if bcat:
            bonus.append({"category": bcat, "keybind": bkey})

    # bar/slot per category, from the Master Sheet
    ws = wb["Master Sheet"]
    barslot = {}
    for r in range(2, ws.max_row + 1):
        cat = cell(ws.cell(r, 4).value)
        if cat:
            barslot[cat] = (cell(ws.cell(r, 1).value), cell(ws.cell(r, 2).value))
    for b in buckets:
        bs = barslot.get(b["category"])
        if bs:
            b["bar"], b["slot"] = bs

    # 2. per-spec ability map, read from each class tab (authoritative headers)
    specs = []
    for cl in CLASS_TABS:
        ws = wb[cl]
        spec_cols = []
        for c in range(3, ws.max_column + 1):
            h = cell(ws.cell(1, c).value)
            if h in (None, "Notes", "Legend"):
                break
            spec_cols.append((c, h))
        for c, spec in spec_cols:
            abilities = {}
            for r in range(2, ws.max_row + 1):
                cat, a = cell(ws.cell(r, 1).value), cell(ws.cell(r, c).value)
                if cat and a:
                    abilities[cat] = a
            specs.append({"class": cl, "spec": spec, "abilities": abilities})

    # 3. reference tables
    def table(name):
        ws = wb[name]
        hdr = [cell(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {
                hdr[c - 1]: cell(ws.cell(r, c).value)
                for c in range(1, ws.max_column + 1)
                if hdr[c - 1]
            }
            if any(row.values()):
                rows.append(row)
        return rows

    return {
        "source": "Bellular Midnight Keybind Planner (Oldschool Edition) — user copy",
        "notation": {
            "digits": "number row (1-4)", "letters": "Q E R F etc",
            "S<key>": "Shift+key", "C<key>": "Ctrl+key", "A<key>": "Alt+key",
            "F1-F4": "stance/form bars",
        },
        "buckets": buckets,
        "bonus_binds": bonus,
        "specs": specs,
        "items": table("Items"),
        "buffs": table("Buffs"),
    }


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT_XLSX)
    if not xlsx.exists():
        sys.exit(f"source workbook not found: {xlsx}")
    data = extract(xlsx)
    JSON_OUT.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n")
    n_ab = sum(len(s["abilities"]) for s in data["specs"])
    print(f"wrote archive side file {JSON_OUT.relative_to(PROJECT)}")
    print(
        f"  buckets={len(data['buckets'])} specs={len(data['specs'])} "
        f"abilities={n_ab} items={len(data['items'])} buffs={len(data['buffs'])}"
    )
    print(
        "NOTE: this did NOT touch the canonical seed or Data.lua. Diff against\n"
        "  data/bellular-keybinds.seed.json by hand; regenerate Lua via\n"
        "  tool/gen_data_lua.py."
    )


if __name__ == "__main__":
    main()
